from langdetect import detect
import openpyxl
from openpyxl.workbook import Workbook


def get_summaries(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    _summaries = []
    for row in ws.iter_rows():
        for cell in row:
            _summaries.append(cell.value)
    return _summaries


def separate_tr_en(_summaries):
    _tr_summaries = []
    _en_summaries = []
    for summary in _summaries:
        try:
            if "[TR] " in summary and "[EN] " in summary:
                print("TREN")
                index_tr = summary.find("[TR] ")
                index_en = summary.find("[EN] ")
                if index_tr < index_en:
                    parts = summary.split('[TR] ')[1].split('[EN]')
                    turkish_part = parts[0].strip()
                    english_part = parts[1].strip()
                else:
                    parts = summary.split('[EN]')[1].split('[TR]')
                    english_part = parts[0].strip()
                    turkish_part = parts[1].strip()
                lang = detect(turkish_part)
                if lang == "tr":
                    _tr_summaries.append(turkish_part)
                    _en_summaries.append(english_part)
                elif lang == "en":
                    _tr_summaries.append(english_part)
                    _en_summaries.append(turkish_part)
                else:
                    print("Türkçe veya İngilizce olmayan metin saptandı. Dil kodu: ", lang)
            elif "[TR] " in summary:
                print("TR")
                turkish_part = summary.removeprefix("[TR] ")
                lang = detect(turkish_part)
                if lang == "tr":
                    _tr_summaries.append(turkish_part)
                elif lang == "en":
                    _en_summaries.append(turkish_part)
                else:
                    print("Türkçe veya İngilizce olmayan metin saptandı. Dil kodu: ", lang)
            elif "[EN] " in summary:
                print("EN")
                english_part = summary.removeprefix("[EN] ")
                lang = detect(english_part)
                if lang == "en":
                    _en_summaries.append(english_part)
                elif lang == "tr":
                    _tr_summaries.append(english_part)
                else:
                    print("Türkçe veya İngilizce olmayan metin saptandı. Dil kodu: ", lang)
        except TypeError as e:
            print(e)
    return _tr_summaries, _en_summaries


def export_summaries(path, _summaries):
    wb = Workbook()
    ws = wb.active
    for summary in _summaries:
        ws.append((summary, ))
    wb.save(path)
    wb.close()


summaries = get_summaries("data/deep_learning_summaries.xlsx")
data = separate_tr_en(summaries)
tr_summaries = data[0]
en_summaries = data[1]
export_summaries("data/tr_summaries.xlsx", tr_summaries)
export_summaries("data/en_summaries.xlsx", en_summaries)
